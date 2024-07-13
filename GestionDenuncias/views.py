from django.shortcuts import render
from django.shortcuts import redirect
from datetime import datetime, timedelta
from django.db.models import Count
from django.db import connection
from .models import Denuncias, Adjuntos, Abogados, Ire, Aportes, Cartola, Formulariosig, EncargadosRegionales, WorkflowActas
from .forms import *
from .forms import GestionTerreno
from openpyxl import Workbook
import csv
from django.contrib.auth.models import User
from django.db.models import Q
from .models import ActasTerreno, ActasRemotas, Tokens, RevisoresDR, EFRDR
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist
#from django.http import HttpResponse
#from django.conf import settings
#from django.core.files.storage import FileSystemStorage
#import os
#import xlrd
import re
from django.conf import settings
from django.views import View

from django.http import JsonResponse
import json
from django.shortcuts import get_object_or_404

#Views Administrador

def denuncias_ingreso(request):
    if request.method == 'POST':
        form = DenunciasForm(request.POST, request.FILES)
        files = request.FILES.getlist('adjunto_denuncia')
        if form.is_valid():
            instancia = form.save(commit=False)
            form.save()
            #for f in files:
            #    file_instance = Adjuntos(id_denuncia=instancia, archivos=f, tipo='adjunto_denuncia')
            #    file_instance.save()
            return redirect('denuncias_ingreso')
    else:
        form = DenunciasForm()

    lista_abogados = Abogados.objects.filter(habilitado=True)


    context = {'form': form, 'lista_abogados': lista_abogados}
    return render(request, 'GestionDenuncias/admin_ingreso_individual.html', context)

def denuncias_ingreso_mass(request):
    message = ''
    if request.method == 'POST':
        form = UpdateDetailsForm(request.POST, request.FILES)
        if form.is_valid():
            # import your django model here like from django.appname.models import model_name
            excel_file = request.FILES['excel_file']
            import os
            import tempfile
            import xlrd
            fd, path = tempfile.mkstemp()  # mkstemp returns a tuple: an integer (index) called file descriptor used by OS to refer to a file and its path
            try:
                with os.fdopen(fd, 'wb') as tmp:
                    tmp.write(excel_file.read())
                book = xlrd.open_workbook(path)
                sheet = book.sheet_by_index(0)
                for rx in range(1, sheet.nrows):
                    obj = Denuncias(
                        numero=sheet.cell_value(rowx=rx, colx=0),
                        link_adjuntos=sheet.cell_value(rowx=rx, colx=1),
                        fecha_ingreso=sheet.cell_value(rowx=rx, colx=2),
                        via_de_ingreso=sheet.cell_value(rowx=rx, colx=3),
                        nombre_denunciante=sheet.cell_value(rowx=rx, colx=4),
                        nombre_denunciado=sheet.cell_value(rowx=rx, colx=5),
                        obs_ingreso=sheet.cell_value(rowx=rx, colx=6),
                        abogado_asistente_id=sheet.cell_value(rowx=rx, colx=7),
                        asignacion_dr_id=sheet.cell_value(rowx=rx, colx=8),
                    )
                    obj.save()
            finally:
                os.remove(path)
                Denuncias.objects.filter(asignacion_dr_id='SIN').update(asignacion_dr_id=None)
        else:
            message = 'Invalid Entries'
    else:
        form = UpdateDetailsForm()

    lista_abogados = Abogados.objects.filter(habilitado=True)

    context = {'form': form, 'message': message, 'lista_abogados': lista_abogados}
    return render(request, 'GestionDenuncias/ingreso_masivo.html', context)

    #return render(request, 'GestionDenuncias/ingreso_masivo.html', context)

def admin_despacho(request):

    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="DESPACHO")
    context = {'todasdenuncias': denuncia_obj_3}
    return render(request, 'GestionDenuncias/admin_despacho.html', context)

def denuncias_enviadas_ad(request):

    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.all
    context = {'todasdenuncias': denuncia_obj_3}
    return render(request, 'GestionDenuncias/admin_bandeja_ingresados.html', context)


# Views Abogado

def abogado_inicio(request):

    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="ENVIADO_JEFE")
    encargados = EncargadosRegionales.objects.all()
    context = {'todasdenuncias': denuncia_obj_3, 'encargados':encargados}
    return render(request, 'GestionDenuncias/abogado_inicio.html', context)

def abogado_gestiones(request):
    adjuntos_obj = Adjuntos.objects.filter(tipo__icontains="adjunto_denuncia")
    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="ENVIADO_JEFE")
    context = {'todasdenuncias': denuncia_obj_3, 'todosadjuntos': adjuntos_obj}
    return render(request, 'GestionDenuncias/abogado_gestiones.html', context)

def abogado_evaluacion_dr(request):

    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.exclude(asignacion_dr=None)
    context = {'todasdenuncias': denuncia_obj_3}
    return render(request, 'GestionDenuncias/abogado_evaluacion_dr.html', context)



def abogado_evaluacion_dr_ver(request, id_denuncia):

    instance = get_object_or_404(Denuncias, id=id_denuncia)
    denuncia_obj_4 = Denuncias.objects.filter(id=id_denuncia)
    form = VerFiscalizacionDR(request.POST or None, instance=instance)
    context = {'todasdenuncias': denuncia_obj_4, 'form': form}
    return render(request, 'GestionDenuncias/abogado_resultado_fiscalizacion.html', context)


def abogado_evaluacion(request):

    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(Q(estado_jefe__iexact="INGRESO")|Q(estado_jefe__icontains="FISCALIZADO_DR")|Q(estado_jefe__icontains="EVALUADO_DR_NO_POSIBLE_FISCALIZAR"))
    context = {'todasdenuncias': denuncia_obj_3}
    return render(request, 'GestionDenuncias/abogado_evaluacion.html', context)

def abogado_rechazos(request):
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="DEVUELTO_JEFE")
    context = {'todasdenuncias': denuncia_obj_3}
    return render(request, 'GestionDenuncias/abogado_rechazos.html', context)

def abogado_comprobacion(request):
    adjuntos_obj = Adjuntos.objects.filter(tipo__icontains="adjunto_denuncia")
    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="INGRESO")
    context = {'todasdenuncias': denuncia_obj_3, 'todosadjuntos': adjuntos_obj}
    return render(request, 'GestionDenuncias/abogado_comprobacion.html', context)

def gestion_denuncia_comp(request, id_denuncia):
    instance = get_object_or_404(Denuncias, id=id_denuncia)
    denuncia_obj_4 = Denuncias.objects.filter(id=id_denuncia)
    form = CompruebaDenuncia(request.POST or None, instance=instance)
    form2 = DetallesDenuncia(request.POST or None, instance=instance)

    context = {'todasdenuncias': denuncia_obj_4, 'form': form, 'form2': form2}

    if request.method == 'POST':

            if request.POST.get(str("guardac")) == 'SI':
                Denuncias.objects.filter(id=str(id_denuncia)).update(estado_jefe="ACTIVADA_COMPROBADA_ABOGADO")
                Denuncias.objects.filter(id=str(id_denuncia)).update(fecha_comprobacion_abogado=datetime.now())
                Denuncias.objects.filter(id=str(id_denuncia)).update(resultado_comprobacion=request.POST.get(str("resultado_comprobacion")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(obs_abogado=request.POST.get(str("obs_abogado")))
                return redirect("abogado_comprobacion")
            if request.POST.get(str("guardac")) == 'NO':
                Denuncias.objects.filter(id=str(id_denuncia)).update(resultado_comprobacion=request.POST.get(str("resultado_comprobacion")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(obs_abogado=request.POST.get(str("obs_abogado")))
                return redirect("abogado_comprobacion")

    return render(request, 'GestionDenuncias/abogado_gestionar_denuncia_comprob.html', context)

def abogado_gestion_denuncia_ac(request, id_denuncia):

    instance = get_object_or_404(Denuncias, id=id_denuncia)
    denuncia_obj_4 = Denuncias.objects.filter(id=id_denuncia)
    form = ActivaDenuncia(request.POST or None, instance=instance)
    form2 = DetallesDenuncia(request.POST or None, instance=instance)

    context = {'todasdenuncias': denuncia_obj_4, 'form': form, 'form2': form2}

    if request.method == 'POST':

            if form.is_valid():
                form.save()
            if request.POST.get(str("guarda")) == 'SI':
                Denuncias.objects.filter(id=str(id_denuncia)).update(estado_jefe="GEST_INGRESO_ABOGADO_REALIZADA")
                Denuncias.objects.filter(id=str(id_denuncia)).update(fecha_evaluacion_abogado=datetime.now())
                return redirect("abogado_evaluacion")
            if request.POST.get(str("guarda")) == 'NO':
                return redirect("abogado_evaluacion")
    return render(request, 'GestionDenuncias/abogado_gestionar_denuncia_activar.html', context)

def abogado_gestion_denuncia_desac(request, id_denuncia):

    instance = get_object_or_404(Denuncias, id=id_denuncia)
    denuncia_obj_4 = Denuncias.objects.filter(id=id_denuncia)
    form = DesactivaDenuncia(request.POST or None, instance=instance)
    form2 = DetallesDenuncia(request.POST or None, instance=instance)
    Denuncias.objects.filter(id=str(id_denuncia)).update(guarda="NO")

    context = {'todasdenuncias': denuncia_obj_4, 'form': form, 'form2': form2}

    if request.method == 'POST':

            if form.is_valid():
                form.save()
            if request.POST.get(str("guarda")) == 'SI':
                Denuncias.objects.filter(id=str(id_denuncia)).update(estado_jefe="DESACTIVADO_ENVIADO_ABOGADO")
                Denuncias.objects.filter(id=str(id_denuncia)).update(fecha_evaluacion_abogado=datetime.now())
                return redirect("abogado_evaluacion")
            if request.POST.get(str("guarda")) == 'NO':
                return redirect("abogado_evaluacion")

    return render(request, 'GestionDenuncias/abogado_gestionar_denuncia_desactivar.html', context)

def abogado_enviados(request):
        # Aca en icontains pongo el filtro con el metodo icontains que es un like
        denuncia_obj_3 = Denuncias.objects.exclude(estado_jefe__icontains="INGRESO").exclude(estado_jefe__icontains="GEST_INGRESO_ABOGADO_REALIZADA")
        context = {'todasdenuncias': denuncia_obj_3}
        return render(request, 'GestionDenuncias/abogado_enviados.html', context)


#Views DR

def dr_evaluacion(request):
    encargados = EncargadosRegionales.objects.filter(id_usuario__username=request.user.username).first()
    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__iexact="INGRESO", asignacion_dr=encargados.dr_asignada_id)
    context = {'todasdenuncias': denuncia_obj_3,'encargados':encargados}
    return render(request, 'GestionDenuncias/dr_evaluacion.html', context)

def dr_fiscalizacion(request):
    encargados = EncargadosRegionales.objects.filter(id_usuario__username=request.user.username).first()
    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="EVALUADO_DR_POSIBLE_FISCALIZAR", asignacion_dr=encargados.dr_asignada_id)
    context = {'todasdenuncias': denuncia_obj_3,'encargados':encargados}
    return render(request, 'GestionDenuncias/dr_fiscalizacion.html', context)


def dr_resultadofiscalizacion(request, id_denuncia):
    encargados = EncargadosRegionales.objects.filter(id_usuario__username=request.user.username).first()
    instance = get_object_or_404(Denuncias, id=id_denuncia)
    denuncia_obj_4 = Denuncias.objects.filter(id=id_denuncia)
    form = FiscalizacionDR(request.POST or None, instance=instance)
    form2 = DetallesDenuncia(request.POST or None, instance=instance)

    context = {'todasdenuncias': denuncia_obj_4, 'form': form, 'form2': form2,'encargados':encargados}

    if request.method == 'POST':

            if request.POST.get(str("dr_guardac")) == 'SI':
                Denuncias.objects.filter(id=str(id_denuncia)).update(estado_jefe="FISCALIZADO_DR")
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_fecha_fiscalizacion=datetime.now())
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_id_inspeccion_survey=request.POST.get(str("dr_id_inspeccion_survey")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_link_carpeta_fiscalizacion=request.POST.get(str("dr_link_carpeta_fiscalizacion")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_nro_requerimiento_candidato=request.POST.get(str("dr_nro_requerimiento_candidato")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(motivo_dr=request.POST.get(str("motivo_dr")))
                if request.POST.get(str("dr_fecha_requerimiento_candidato")) != '':
                    Denuncias.objects.filter(id=str(id_denuncia)).update(dr_fecha_requerimiento_candidato=request.POST.get(str("dr_fecha_requerimiento_candidato")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_resultado_requerimiento_candidato=request.POST.get(str("dr_resultado_requerimiento_candidato")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_retiro_municipio=request.POST.get(str("dr_retiro_municipio")))
                if request.POST.get(str("dr_fecha_retiro_municipio")) != '':
                    Denuncias.objects.filter(id=str(id_denuncia)).update(dr_fecha_retiro_municipio=request.POST.get(str("dr_fecha_retiro_municipio")))
                return redirect("dr_fiscalizacion")

            if request.POST.get(str("dr_guardac")) == 'NO':
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_id_inspeccion_survey=request.POST.get(str("dr_id_inspeccion_survey")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_link_carpeta_fiscalizacion=request.POST.get(str("dr_link_carpeta_fiscalizacion")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_nro_requerimiento_candidato=request.POST.get(str("dr_nro_requerimiento_candidato")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(motivo_dr=request.POST.get(str("motivo_dr")))
                if request.POST.get(str("dr_fecha_requerimiento_candidato")) != '':
                    Denuncias.objects.filter(id=str(id_denuncia)).update(dr_fecha_requerimiento_candidato=request.POST.get(str("dr_fecha_requerimiento_candidato")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_resultado_requerimiento_candidato=request.POST.get(str("dr_resultado_requerimiento_candidato")))
                Denuncias.objects.filter(id=str(id_denuncia)).update(dr_retiro_municipio=request.POST.get(str("dr_retiro_municipio")))
                if request.POST.get(str("dr_fecha_retiro_municipio")) != '':
                    Denuncias.objects.filter(id=str(id_denuncia)).update(dr_fecha_retiro_municipio=request.POST.get(str("dr_fecha_retiro_municipio")))
                return redirect("dr_fiscalizacion")

    return render(request, 'GestionDenuncias/dr_resultado_fiscalizacion.html', context)

def dr_enviados(request):
    encargados = EncargadosRegionales.objects.filter(id_usuario__username=request.user.username).first()
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(asignacion_dr=encargados.dr_asignada_id).exclude(estado_jefe__iexact="INGRESO")
    context = {'todasdenuncias': denuncia_obj_3,'encargados':encargados}
    return render(request, 'GestionDenuncias/dr_enviados.html', context)

def dr_evaluacion_dr_ver(request, id_denuncia):

    instance = get_object_or_404(Denuncias, id=id_denuncia)
    denuncia_obj_4 = Denuncias.objects.filter(id=id_denuncia)
    form = VerFiscalizacionDR(request.POST or None, instance=instance)
    context = {'todasdenuncias': denuncia_obj_4, 'form': form}
    return render(request, 'GestionDenuncias/dr_ver_resultado_fiscalizacion.html', context)


#Views Jefe

def jefe_inicio(request):
    return render(request, 'GestionDenuncias/jefe_inicio.html')

def jefe_evaluacion_act(request):
    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="ACTIVADA_COMPROBADA_ABOGADO")
    context = {'todasdenuncias': denuncia_obj_3}
    return render(request, 'GestionDenuncias/jefe_evaluacion_act.html', context)

def jefe_evaluacion_desact(request):
    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="DESACTIVADO_ENVIADO_ABOGADO")
    context = {'todasdenuncias': denuncia_obj_3}
    return render(request, 'GestionDenuncias/jefe_evaluacion_desact.html', context)

def validacion_masiva(request):
    #Aca en icontains pongo el filtro con el metodo icontains que es un like
    denuncia_obj_3 = Denuncias.objects.filter(estado_jefe__icontains="DESACTIVADO_ENVIADO_ABOGADO")
    codigos = Denuncias.objects.values('codigo_desactivacion') \
  .annotate(cantidad=Count('numero')) \
        .filter(estado_jefe__icontains='DESACTIVADO_ENVIADO_ABOGADO')

    context = {'todasdenuncias': denuncia_obj_3, 'codigos': codigos}
    return render(request, 'GestionDenuncias/jefe_validacion_masiva.html', context)

def modifica_denuncia(request):
    data = json.loads(request.body)

    if data['datos']['tipo'] == 'rechaza':
         Denuncias.objects.filter(id=str(data['datos']['id_denuncia'])).update(estado_jefe='DEVUELTO_JEFE')
         Denuncias.objects.filter(id=str(data['datos']['id_denuncia'])).update(obs_jefe=data['datos']['motivo_rechazo'])

    if data['datos']['tipo'] == 'acepta':
         Denuncias.objects.filter(id=str(data['datos']['id_denuncia'])).update(obs_jefe='ACEPTADO')
         if data['datos']['categoria'] == 'activa':
            Denuncias.objects.filter(id=str(data['datos']['id_denuncia'])).update(estado_jefe='ACTIVADO_DESPACHO')
         if data['datos']['categoria'] == 'desactiva':
            Denuncias.objects.filter(id=str(data['datos']['id_denuncia'])).update(estado_jefe='DESACTIVADO_DESPACHO')

    if data['datos']['tipo'] == 'acepta_masiva':
         Denuncias.objects.filter(codigo_desactivacion=str(data['datos']['id_denuncia'])).update(obs_jefe='ACEPTADO')
         Denuncias.objects.filter(codigo_desactivacion=str(data['datos']['id_denuncia'])).update(estado_jefe='DESACTIVADO_DESPACHO')
    if data['datos']['tipo'] == 'rechaza_masiva':
         Denuncias.objects.filter(codigo_desactivacion=str(data['datos']['id_denuncia'])).update(estado_jefe='DEVUELTO_JEFE')
         Denuncias.objects.filter(codigo_desactivacion=str(data['datos']['id_denuncia'])).update(obs_jefe=data['datos']['motivo_rechazo'])

    if data['datos']['tipo'] == 'esposible':
         Denuncias.objects.filter(id=str(data['datos']['id_denuncia'])).update(estado_jefe='EVALUADO_DR_POSIBLE_FISCALIZAR')

    if data['datos']['tipo'] == 'noesposible':
         Denuncias.objects.filter(id=str(data['datos']['id_denuncia'])).update(estado_jefe='EVALUADO_DR_NO_POSIBLE_FISCALIZAR')
         Denuncias.objects.filter(id=str(data['datos']['id_denuncia'])).update(motivo_dr=data['datos']['motivo_rechazo'])

    return JsonResponse([str(data['datos']['id_denuncia']), 'DEVUELTO_JEFE'], safe=False)


#Views SAR
def auditor_aportes(request, rut):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    aportes = Aportes.objects.filter(rut_receptor_id=rut)
    context = {'aportes': aportes}
    return render(request, 'GestionDenuncias/auditor_aportes.html', context)

def auditor_ire(request, eleccion):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    if eleccion == 1:
        seleccionado = 'ALCALDES'
    if eleccion == 2:
        seleccionado = 'CONCEJALES'
    if eleccion == 3:
        seleccionado = 'CONVENCIONALES CONSTITUYENTES'
    if eleccion == 4:
        seleccionado = 'CONVENCIONALES CONSTITUYENTES - PUEBLOS INDIGENAS'
    if eleccion == 5:
        seleccionado = 'GOBERNADOR REGIONAL'
    if eleccion == 6:
        seleccionado = 'PARTIDO'
    ire = Ire.objects.filter(eleccion=seleccionado)
    context = {'ire': ire}
    return render(request, 'GestionDenuncias/auditor_ire.html', context)

def auditor_bandeja_asignados(request):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    ire = Ire.objects.filter(celula_asignada__isnull=False)
    context = {'ire': ire}
    return render(request, 'GestionDenuncias/auditor_bandeja_asignadas.html', context)

def auditor_cartola(request, rut):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    cartola = Cartola.objects.filter(rut_id=rut)
    context = {'cartola': cartola}
    return render(request, 'GestionDenuncias/auditor_cartola.html', context)

def auditor_avance_general(request):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    context = {}
    return render(request, 'GestionDenuncias/auditor_avance_general.html', context)

def auditor_avance_celula(request):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    context = {}
    return render(request, 'GestionDenuncias/auditor_avance_celula.html', context)

def auditor_candidato(request, rut):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    ire = Ire.objects.filter(rut=rut)
    context = {'ire': ire}
    return render(request, 'GestionDenuncias/auditor_candidato.html', context)

def auditor_indicadores(request):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    context = {}
    return render(request, 'GestionDenuncias/auditor_indicadores.html', context)

def auditor_f87(request, rut):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    f87 = Formulariosig.objects.filter(rut_partido_candidato_id=rut, tpo_rendicion_codigo = 'F87')
    context = {'f87': f87}
    return render(request, 'GestionDenuncias/auditor_f87.html', context)

def auditor_f88(request, rut):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    f88 = Formulariosig.objects.filter(rut_partido_candidato_id=rut, tpo_rendicion_codigo = 'F88')
    context = {'f88': f88}
    return render(request, 'GestionDenuncias/auditor_f88.html', context)


def modifica_candidato(request):
    data = json.loads(request.body)

    if data['datos']['tipo'] == 'comentario':
         objeto = Ire.objects.filter(rut=str(data['datos']['id_candidato'])).get()
         comentario = objeto.comentarios
         if comentario != None:
            comentario = comentario + "\n" + str(data['datos']['username']) + ": " + str(data['datos']['comentario_ingresado'])
         else:
            comentario = str(data['datos']['username']) + ": " + str(data['datos']['comentario_ingresado'])
         Ire.objects.filter(rut=str(data['datos']['id_candidato'])).update(comentarios=str(comentario))

    return JsonResponse([str(data['datos']['id_candidato'])], safe=False)

def numero_a_romano(num):
    # Diccionario de números arábigos a romanos incluyendo el caso especial para el 13
    roman_dict = {
        '1': 'I', '2': 'II', '3': 'III', '4': 'IV', '5': 'V', '6': 'VI',
        '7': 'VII', '8': 'VIII', '9': 'IX', '10': 'X', '11': 'XI', '12': 'XII',
        '13': 'RM', '14': 'XIV', '15': 'XV', '16': 'XVI'
    }
    return roman_dict.get(num, "")

@csrf_exempt
def carga_datos_actas_terreno(request):
    if request.method == 'POST':
        # Extrae los datos de la solicitud POST
        token = request.POST.get('token')
        object_id = request.POST.get('object_id')
        global_id = request.POST.get('global_id')
        fecha = request.POST.get('fecha')
        region_inicial = request.POST.get('region')
        try:
            region = numero_a_romano(region_inicial)
        except:
            region = region_inicial
        ubicacion = request.POST.get('ubicacion')
        comuna = request.POST.get('comuna')
        seleccion_motivo_inspeccion = request.POST.get('seleccion_motivo_inspeccion')
        indique_folio = request.POST.get('indique_folio')
        indique_otro = request.POST.get('indique_otro')
        Sujeto_fiscalizado = request.POST.get('Sujeto_fiscalizado')
        #partido_politico_habilitado = request.POST.get('partido_politico_habilitado')
        #opcion_plebiscitaria = request.POST.get('opcion_plebiscitaria')
        alcalde= request.POST.get('alcalde')
        gore= request.POST.get('gore')
        cuenta_con_104= request.POST.get('cuenta_con_104')
        materia_fiscalizada = request.POST.get('materia_fiscalizada')
        corresponde_espacio_publico_autorizado = request.POST.get('corresponde_espacio_publico_autorizado')
        seleccione_espacio = request.POST.get('seleccione_espacio')
        adosada_bien_nacional = request.POST.get('adosada_bien_nacional')
        nombre_bienes = request.POST.get('nombre_bienes')
        cantidad_elementos_propaganda_publico = request.POST.get('cantidad_elementos_propaganda_publico')
        propaganda_excede_dimensiones = request.POST.get('propaganda_excede_dimensiones')
        seleccione_lugar = request.POST.get('seleccione_lugar')
        otro_antecente = request.POST.get('otro_antecente')
        creation_date = request.POST.get('creation_date')
        creator = request.POST.get('creator')
        edit_date = request.POST.get('edit_date')
        editor = request.POST.get('editor')
        x_coord = request.POST.get('x_coord')
        y_coord = request.POST.get('y_coord')
        evidencia_fotografica = request.POST.get('evidencia_fotografica')
        link_firma_cargo_timbre = request.POST.get('link_firma_cargo_timbre')
        id_inspeccion = request.POST.get('id_inspeccion')
        existe_despliegue_propaganda = request.POST.get('existe_despliegue_propaganda')
        otro_sujeto_fiscalizado = request.POST.get('otro_sujeto_fiscalizado')
        actividad_fiscalizada = request.POST.get('actividad_fiscalizada')
        indique_cantidad_brigadistas_lugar = request.POST.get('indique_cantidad_brigadistas_lugar')

        # Crea una instancia de tu modelo de datos y asigna los valores de la solicitud POST
        acta_terreno = ActasTerreno(
            object_id=int(object_id) + 333,
            global_id=global_id.replace('{','').replace('}',''),
            fecha=fecha,
            region=region,
            ubicacion=ubicacion,
            comuna=comuna,
            seleccion_motivo_inspeccion=seleccion_motivo_inspeccion,
            indique_folio=indique_folio,
            indique_otro=indique_otro,
            Sujeto_fiscalizado=Sujeto_fiscalizado,
            #partido_politico_habilitado=partido_politico_habilitado,
            #opcion_plebiscitaria=opcion_plebiscitaria,
            alcalde=alcalde,
            gore=gore,
            cuenta_con_104 = cuenta_con_104,
            materia_fiscalizada=materia_fiscalizada,
            corresponde_espacio_publico_autorizado=corresponde_espacio_publico_autorizado,
            seleccione_espacio=seleccione_espacio,
            adosada_bien_nacional=adosada_bien_nacional,
            nombre_bienes=nombre_bienes,
            cantidad_elementos_propaganda_publico=cantidad_elementos_propaganda_publico,
            propaganda_excede_dimensiones=propaganda_excede_dimensiones,
            seleccione_lugar=seleccione_lugar,
            otro_antecente=otro_antecente,
            creation_date=creation_date,
            creator=creator,
            edit_date=edit_date,
            editor=editor,
            x_coord=x_coord,
            y_coord=y_coord,
            evidencia_fotografica=evidencia_fotografica,
            link_firma_cargo_timbre=link_firma_cargo_timbre,
            id_inspeccion=id_inspeccion,
            existe_despliegue_propaganda=existe_despliegue_propaganda,
            otro_sujeto_fiscalizado=otro_sujeto_fiscalizado,
            actividad_fiscalizada=actividad_fiscalizada,
            indique_cantidad_brigadistas_lugar=indique_cantidad_brigadistas_lugar
        )

        try:
            objeto_token = Tokens(Token=token, Fecha=datetime.now())
            objeto_token.save()
        except Exception as e:
            print("Error Token:", e)

        # Guarda la instancia en la base de datos
        acta_terreno.save()

        # Retorna una respuesta HTTP 200
        return HttpResponse('Datos guardados correctamente')
    else:
        # Retorna una respuesta HTTP 405 si se recibe una solicitud que no es POST
        return HttpResponse(status=405)


@csrf_exempt
def carga_datos_actas_remotas(request):
    if request.method == 'POST':
        # Extrae los datos de la solicitud POST
        acta_remota_data = {
            'object_id': request.POST.get('object_id'),
            'global_id': request.POST.get('global_id').replace('{','').replace('}',''),
            'fecha': request.POST.get('fecha'),
            'region': numero_a_romano(request.POST.get('region')),
            'seleccion_motivo_inspeccion': request.POST.get('seleccion_motivo_inspeccion'),
            'indique_folio': request.POST.get('indique_folio'),
            'indique_otro': request.POST.get('indique_otro'),
            'sujeto_fiscalizado': request.POST.get('sujeto_fiscalizado'),
            #'partido_politico_habilitado': request.POST.get('partido_politico_habilitado'),
            'otro_sujeto_fiscalizado': request.POST.get('otro_sujeto_fiscalizado'),
            #'opcion_plebiscitaria': request.POST.get('opcion_plebiscitaria'),
            'alcalde': request.POST.get('alcalde'),
            'gore': request.POST.get('gore'),
            'es_medio_pagado': request.POST.get('es_medio_pagado'),
            'medio_fiscalizado': request.POST.get('medio_fiscalizado'),
            'nombre_medio': request.POST.get('nombre_medio'),
            'soporte_material_link': request.POST.get('soporte_material_link'),
            'medio_tiene_tarifario': request.POST.get('medio_tiene_tarifario'),
            'radiofrecuencia_medio': request.POST.get('radiofrecuencia_medio'),
            'rrss_fiscalizada': request.POST.get('rrss_fiscalizada'),
            'usuario_perfil_rrss': request.POST.get('usuario_perfil_rrss'),
            'corresponde_medio_prensa': request.POST.get('corresponde_medio_prensa'),
            'otro_antecente': request.POST.get('otro_antecente'),
            'medios_respaldo_adjunto': request.POST.get('medios_respaldo_adjunto'),
            'ingrese_audios': request.POST.get('ingrese_audios'),
            'link_firma_cargo_timbre': request.POST.get('link_firma_cargo_timbre'),
            'id_inspeccion': request.POST.get('id_inspeccion'),
            'id_workforce': request.POST.get('id_workforce'),
            'creation_date': request.POST.get('creation_date'),
            'creator': request.POST.get('creator'),
            'edit_date': request.POST.get('edit_date'),
            'editor': request.POST.get('editor'),
        }

        # Crea una instancia de tu modelo de datos y asigna los valores
        acta_remota = ActasRemotas(**acta_remota_data)

        #try:
        #    tok = request.POST.get('token')
        #    objeto_token = Tokens(Token=tok, Fecha=datetime.now())
        #    objeto_token.save()
        #except:
        #    print("Error Token")

        # Guarda la instancia en la base de datos
        acta_remota.save()

        # Retorna una respuesta HTTP 200
        return HttpResponse('Datos guardados correctamente')
    else:
        # Retorna una respuesta HTTP 405 si se recibe una solicitud que no es POST
        return HttpResponse(status=405)


def terreno_pendiente_clasificacion(request):

    try:
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_terreno = ActasTerreno.objects.filter(sis_clasificacion="Pendiente", region=region_usuario)
    else:
        actas_terreno = []
    # Aca en icontains pongo el filtro con el metodo icontains que es un like

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Revisor_Pendiente_Clasificacion.html', context)





def increment_url_numbers(url):
    # Esta función toma un match object, incrementa el número y devuelve una cadena
    def replacer(m):
        return str(int(m.group(1)) + 1)

    # Usa una expresión regular para identificar y modificar los números al final de la URL
    try:
        new_url = re.sub(r'(\d+)$', replacer, url)
    except:
        return 'Http://'
    return new_url


def pasar_acta(request):
    data = json.loads(request.body)
    if data['datos']['etapa'] == 'archivo_terreno':
         ActasTerreno.objects.filter(global_id=str(data['datos']['global_id'])).update(sis_clasificacion=str(data['datos']['etapa']),
                                                                                       sis_motivo_inicial = str(data['datos']['inputmotivoValue']),
                                                                                       fecha_clasificacion_revisor = datetime.now())
         #WorkflowSCP.objects.create(rut_candidato_partido=str(data['datos']['rut']), usuario=str(request.user),
         #                                       nueva_etapa=str(data['datos']['etapa']),
         #                                       fecha_cambio=datetime.now())
    if data['datos']['etapa'] == 'con_infraccion_revisor_terreno':
        if data['datos']['codigo'] != 'Pendiente':
            ActasTerreno.objects.filter(global_id=str(data['datos']['global_id'])).update(
                sis_clasificacion=str(data['datos']['etapa']), sis_codigo = str(data['datos']['codigo']),
                                                                                       sis_motivo_inicial = str(data['datos']['inputmotivoValue']),
                                                                                       fecha_clasificacion_revisor = datetime.now())
        # WorkflowSCP.objects.create(rut_candidato_partido=str(data['datos']['rut']), usuario=str(request.user),
        #                                       nueva_etapa=str(data['datos']['etapa']),
        #                                       fecha_cambio=datetime.now())

    if data['datos']['etapa'] == 'archivo_remota':
         ActasRemotas.objects.filter(global_id=str(data['datos']['global_id'])).update(sis_clasificacion=str(data['datos']['etapa']),
                                                                                       sis_motivo_inicial = str(data['datos']['inputmotivoValue']),
                                                                                       fecha_clasificacion_revisor = datetime.now())
         #WorkflowSCP.objects.create(rut_candidato_partido=str(data['datos']['rut']), usuario=str(request.user),
         #                                       nueva_etapa=str(data['datos']['etapa']),
         #                                       fecha_cambio=datetime.now())
    if data['datos']['etapa'] == 'con_infraccion_revisor_remota':
        if data['datos']['codigo'] != 'Pendiente':
            ActasRemotas.objects.filter(global_id=str(data['datos']['global_id'])).update(
                sis_clasificacion=str(data['datos']['etapa']), sis_codigo = str(data['datos']['codigo']),
                                                                                       sis_motivo_inicial = str(data['datos']['inputmotivoValue']),
                                                                                       fecha_clasificacion_revisor = datetime.now())
        # WorkflowSCP.objects.create(rut_candidato_partido=str(data['datos']['rut']), usuario=str(request.user),
        #                                       nueva_etapa=str(data['datos']['etapa']),
        #                                       fecha_cambio=datetime.now())
    if data['datos']['etapa'] == 'asignado_Abogado':
        ActasRemotas.objects.filter(global_id=str(data['datos']['global_id'])).update(
                sis_clasificacion=str(data['datos']['etapa']), abogado_asignado = str(data['datos']['asignacion_selected']))
        ActasTerreno.objects.filter(global_id=str(data['datos']['global_id'])).update(
                sis_clasificacion=str(data['datos']['etapa']),  abogado_asignado = str(data['datos']['asignacion_selected']))
        # WorkflowSCP.objects.create(rut_candidato_partido=str(data['datos']['rut']), usuario=str(request.user),
        #                                       nueva_etapa=str(data['datos']['etapa']),
        #                                       fecha_cambio=datetime.now())

    WorkflowActas.objects.create(GlobalID=str(data['datos']['global_id']),
                                     Usuario=request.user.username, NuevaEtapa=str(data['datos']['etapa']),
                                     FechaCambio=datetime.now())

    return JsonResponse([str(data['datos']['global_id']), 'Asignado'], safe=False)

def remota_pendiente_clasificacion(request):
    # Aca en icontains pongo el filtro con el metodo icontains que es un like
    try:
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_remota = ActasRemotas.objects.filter(sis_clasificacion="Pendiente", region=region_usuario)
    else:
        actas_remota = []

    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Revisor_Pendiente_Clasificacion.html', context)


def terreno_con_infraccion(request):
    try:
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_terreno = ActasTerreno.objects.filter(sis_clasificacion="con_infraccion_revisor_terreno", region=region_usuario)
    else:
        actas_terreno = []

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Revisor_Con_Infraccion.html', context)

def terreno_archivo(request):
    try:
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_terreno = ActasTerreno.objects.filter(sis_clasificacion="archivo_terreno", region=region_usuario)
    else:
        actas_terreno = []

    # Aca en icontains pongo el filtro con el metodo icontains que es un like

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Revisor_Archivo.html', context)

def remota_archivo(request):
    try:
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_remota = ActasRemotas.objects.filter(sis_clasificacion="archivo_remota", region=region_usuario)
    else:
        actas_remota = []
    # Aca en icontains pongo el filtro con el metodo icontains que es un like

    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Revisor_Archivo.html', context)


def terreno_con_infraccion_gestiones(request, id):
    actas_terreno = ActasTerreno.objects.filter(global_id=id)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    if request.method == 'POST':
        GestionTerrenoForm = GestionTerreno(request.POST, instance=acta)
        if GestionTerrenoForm.is_valid():
            GestionTerrenoForm.save()
            if GestionTerrenoForm.cleaned_data['sis_clasificacion'] == 'EFR_Validacion':
                ActasTerreno.objects.filter(global_id=id).update(
                fecha_envio_efr=datetime.now())
                WorkflowActas.objects.create(GlobalID=str(actas_terreno[0].global_id),
                                             Usuario=request.user.username, NuevaEtapa='EFR_Validacion',
                                             FechaCambio=datetime.now())

            return redirect('terreno_con_infraccion')
    else:
        GestionTerrenoForm = GestionTerreno(instance=acta)



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'GestionTerrenoForm': GestionTerrenoForm}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Revisor_Con_Infraccion_Gestiones.html', context)


def remotas_con_infraccion(request):
    try:
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_remota = ActasRemotas.objects.filter(sis_clasificacion="con_infraccion_revisor_remota", region=region_usuario)
    else:
        actas_remota = []

    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None


    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Revisor_Con_Infraccion.html', context)

def remota_con_infraccion_gestiones(request, id):
    actas_remota = ActasRemotas.objects.filter(global_id=id)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    if request.method == 'POST':
        GestionRemotaForm = GestionRemota(request.POST, instance=acta)


        if GestionRemotaForm.is_valid():
            GestionRemotaForm.save()
            if GestionRemotaForm.cleaned_data['sis_clasificacion'] == 'EFR_Validacion':
                ActasRemotas.objects.filter(global_id=id).update(
                    fecha_envio_efr=datetime.now())
                WorkflowActas.objects.create(GlobalID=str(actas_remota[0].global_id),
                                             Usuario=request.user.username, NuevaEtapa='EFR_Validacion',
                                             FechaCambio=datetime.now()
                                                    )
            return redirect('remotas_con_infraccion')
    else:
        GestionRemotaForm = GestionRemota(instance=acta)


    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'GestionRemotaForm': GestionRemotaForm}
    return render(request, 'GestionDenuncias/SGD2_Remota_Revisor_Con_Infraccion_Gestiones.html', context)


def efr_terreno_con_infraccion(request):
    try:
        revisor = EFRDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except EFRDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_terreno = ActasTerreno.objects.filter(sis_clasificacion="EFR_Validacion", region=region_usuario)
    else:
        actas_terreno = []

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_EFR_Con_Infraccion.html', context)

def efr_remota_con_infraccion(request):
    try:
        revisor = EFRDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except EFRDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_remota = ActasRemotas.objects.filter(sis_clasificacion="EFR_Validacion", region=region_usuario)
    else:
        actas_remota = []

    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    # actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_EFR_Con_Infraccion.html', context)

def efr_terreno_con_infraccion_gestiones(request, id):
    actas_terreno = ActasTerreno.objects.filter(global_id=id)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        if request.method == 'POST':
            GestionTerrenoEFRForm = GestionTerrenoEFR(request.POST, instance=acta)
            if GestionTerrenoEFRForm.is_valid():
                acta = GestionTerrenoEFRForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.sis_resultado_efr = GestionTerrenoEFRForm.cleaned_data['sis_resultado_efr']
                acta.sis_motivo_rechazo = GestionTerrenoEFRForm.cleaned_data['sis_motivo_rechazo']
                acta.sis_clasificacion = GestionTerrenoEFRForm.cleaned_data['sis_clasificacion']
                acta.fecha_envio_ufisca = None
                if GestionTerrenoEFRForm.cleaned_data['sis_clasificacion'] == 'efr_aceptado':
                    acta.fecha_envio_ufisca = datetime.now()
                acta.save(update_fields=['sis_resultado_efr','sis_motivo_rechazo', 'sis_clasificacion', 'fecha_envio_ufisca' ])
                WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                             Usuario=request.user.username, NuevaEtapa=GestionTerrenoEFRForm.cleaned_data['sis_clasificacion'],
                                             FechaCambio=datetime.now())
                return redirect('efr_terreno_con_infraccion')
            else:
                print(GestionTerrenoEFRForm.errors)
                return redirect('efr_terreno_con_infraccion')
        else:
            GestionTerrenoEFRForm = GestionTerrenoEFR(instance=acta)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'GestionTerrenoEFRForm': GestionTerrenoEFRForm}
    return render(request, 'GestionDenuncias/SGD2_Terreno_EFR_Con_Infraccion_Gestiones.html', context)


def efr_remota_con_infraccion_gestiones(request, id):
    actas_remota = ActasRemotas.objects.filter(global_id=id)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        try:
            latest_token = Tokens.objects.latest('id')
        except ObjectDoesNotExist:
            latest_token = None

        if request.method == 'POST':
            GestionRemotasEFRForm = GestionRemotasEFR(request.POST, instance=acta)
            if GestionRemotasEFRForm.is_valid():
                acta = GestionRemotasEFRForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.sis_resultado_efr = GestionRemotasEFRForm.cleaned_data['sis_resultado_efr']
                acta.sis_motivo_rechazo = GestionRemotasEFRForm.cleaned_data['sis_motivo_rechazo']
                acta.sis_clasificacion = GestionRemotasEFRForm.cleaned_data['sis_clasificacion']
                acta.fecha_envio_ufisca = None
                if GestionRemotasEFRForm.cleaned_data['sis_clasificacion'] == 'efr_aceptado':
                    acta.fecha_envio_ufisca = datetime.now()
                acta.save(update_fields=['sis_resultado_efr', 'sis_motivo_rechazo', 'sis_clasificacion', 'fecha_envio_ufisca'])

                WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                             Usuario=request.user.username, NuevaEtapa=GestionRemotasEFRForm.cleaned_data['sis_clasificacion'],
                                             FechaCambio=datetime.now())

                return redirect('efr_remota_con_infraccion')
            else:
                print(GestionRemotasEFRForm.errors)
                return redirect('efr_remota_con_infraccion')
        else:
            GestionRemotasEFRForm = GestionRemotasEFR(instance=acta)

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'GestionRemotasEFRForm': GestionRemotasEFRForm}
    return render(request, 'GestionDenuncias/SGD2_Remota_EFR_Con_Infraccion_Gestiones.html', context)

def terreno_con_infraccion_rechazo(request):
    try:
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_terreno = ActasTerreno.objects.filter(sis_clasificacion="revisor_rechazo", region=region_usuario)
    else:
        actas_terreno = []

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Revisor_Con_Infraccion_Rechazo.html', context)


def terreno_con_infraccion_rechazo_gestiones(request, id):
    actas_terreno = ActasTerreno.objects.filter(global_id=id)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    if request.method == 'POST':
        GestionTerrenoForm = GestionTerreno(request.POST, instance=acta)
        if GestionTerrenoForm.is_valid():
            GestionTerrenoForm.save()

            return redirect('terreno_con_infraccion_rechazo')
    else:
        GestionTerrenoForm = GestionTerreno(instance=acta)



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'GestionTerrenoForm': GestionTerrenoForm}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Revisor_Con_Infraccion_Rechazo_Gestiones.html', context)


def remotas_con_infraccion_rechazo(request):
    try:
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
        print("Reg_ok")
    except RevisoresDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None
        print("Reg_vacia")
    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_remota = ActasRemotas.objects.filter(sis_clasificacion="revisor_rechazo", region=region_usuario)

    else:
        actas_remota = []

    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Revisor_Con_Infraccion_Rechazo.html', context)

def remota_con_infraccion_gestiones_rechazo(request, id):
    actas_remota = ActasRemotas.objects.filter(global_id=id)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    if request.method == 'POST':
        GestionRemotaForm = GestionRemota(request.POST, instance=acta)
        if GestionRemotaForm.is_valid():
            GestionRemotaForm.save()
            WorkflowActas.objects.create(GlobalID=str(actas_remota[0].global_id),
                                         Usuario=request.user.username, NuevaEtapa='EFR_Validacion',
                                         FechaCambio=datetime.now())
            return redirect('remotas_con_infraccion_rechazo')
    else:
        GestionRemotaForm = GestionRemota(instance=acta)


    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'GestionRemotaForm': GestionRemotaForm}
    return render(request, 'GestionDenuncias/SGD2_Remota_Revisor_Con_Infraccion_Gestiones_Rechazo.html', context)


def expcsv(request, lc):
    # Crear la respuesta HTTP como un archivo CSV
    if lc == 'abeced':
        wb = Workbook()
        ws = wb.active

        # Opcionalmente, escribir los nombres de las columnas
        column_names = [field.name for field in ActasTerreno._meta.fields]
        ws.append(column_names)

        # Escribir los datos del modelo
        for obj in ActasTerreno.objects.all():
            ws.append([getattr(obj, field.name) for field in ActasTerreno._meta.fields])

        # Configurar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ActasTerreno.xlsx'

        # Guardar el libro de Excel en la respuesta
        wb.save(response)

        return response



    elif lc == 'obeced':
        wb = Workbook()
        ws = wb.active

        # Opcionalmente, escribir los nombres de las columnas
        column_names = [field.name for field in ActasRemotas._meta.fields]
        ws.append(column_names)

        # Escribir los datos del modelo
        for obj in ActasRemotas.objects.all():
            ws.append([getattr(obj, field.name) for field in ActasRemotas._meta.fields])

        # Configurar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ActasRemotas.xlsx'

        # Guardar el libro de Excel en la respuesta
        wb.save(response)

        return response

    elif lc == 'workflow':
        wb = Workbook()
        ws = wb.active

        # Opcionalmente, escribir los nombres de las columnas
        column_names = [field.name for field in ActasRemotas._meta.fields]
        ws.append(column_names)

        # Escribir los datos del modelo
        for obj in WorkflowActas.objects.all():
            ws.append([getattr(obj, field.name) for field in WorkflowActas._meta.fields])

        # Configurar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=WorkflowActas.xlsx'

        # Guardar el libro de Excel en la respuesta
        wb.save(response)

        return response

    else:
        print("Error")

def admin_terreno_con_infraccion(request):
        # Filtra las ActasTerreno basadas en la región del usuario
    actas_terreno = ActasTerreno.objects.filter(sis_clasificacion="efr_aceptado")
    abogados = Abogados.objects.filter(habilitado=True)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'abogados':abogados}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Admin_Con_Infraccion.html', context)

def admin_remota_con_infraccion(request):

    actas_remota = ActasRemotas.objects.filter(sis_clasificacion="efr_aceptado")
    abogados = Abogados.objects.filter(habilitado=True)

    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    # actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'abogados':abogados}
    return render(request, 'GestionDenuncias/SGD2_Remota_Admin_Con_Infraccion.html', context)

def abogado_evaluacion_terreno(request):

    usuario_actual=request.user.username
    abogado_actual = Abogados.objects.filter(username=usuario_actual)

    actas_terreno = ActasTerreno.objects.filter(Q(sis_clasificacion__iexact="asignado_Abogado")|Q(sis_clasificacion__iexact="rechaza_encargado"), abogado_asignado=abogado_actual.first().rut)

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Abogado_Evaluacion.html', context)

def abogado_activar_terreno(request, id):
    actas_terreno = ActasTerreno.objects.filter(global_id=id)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        if request.method == 'POST':
            GestionTerrenoAbogadoActivarForm = GestionTerrenoAbogadoActivar(request.POST, instance=acta)
            if GestionTerrenoAbogadoActivarForm.is_valid():
                acta = GestionTerrenoAbogadoActivarForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.abogado_resultado = GestionTerrenoAbogadoActivarForm.cleaned_data['abogado_resultado']
                acta.abogado_motivo_devolucion = GestionTerrenoAbogadoActivarForm.cleaned_data['abogado_motivo_devolucion']
                acta.sis_clasificacion = GestionTerrenoAbogadoActivarForm.cleaned_data['sis_clasificacion']
                acta.abogado_eleccion = GestionTerrenoAbogadoActivarForm.cleaned_data['abogado_eleccion']
                acta.abogado_presunto_infractor = GestionTerrenoAbogadoActivarForm.cleaned_data['abogado_presunto_infractor']
                acta.abogado_codigo_activa = GestionTerrenoAbogadoActivarForm.cleaned_data['abogado_codigo_activa']
                acta.abogado_obs = GestionTerrenoAbogadoActivarForm.cleaned_data['abogado_obs']
                acta.fecha_evaluacion_abogado = None
                if GestionTerrenoAbogadoActivarForm.cleaned_data['sis_clasificacion'] == 'abogado_activado':
                    acta.fecha_evaluacion_abogado = datetime.now()
                    WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                                 Usuario=request.user.username,
                                                 NuevaEtapa=GestionTerrenoAbogadoActivarForm.cleaned_data['sis_clasificacion'],
                                                 FechaCambio=datetime.now())
                acta.save(update_fields=['abogado_resultado','abogado_motivo_devolucion', 'sis_clasificacion', 'abogado_eleccion', 'abogado_presunto_infractor' , 'abogado_codigo_activa', 'abogado_obs','fecha_evaluacion_abogado'])
                return redirect('abogado_evaluacion_terreno')
            else:
                print(GestionTerrenoAbogadoActivar.errors)
                return redirect('abogado_evaluacion_terreno')
        else:
            GestionTerrenoAbogadoActivarForm = GestionTerrenoAbogadoActivar(instance=acta)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'GestionTerrenoAbogadoActivarForm': GestionTerrenoAbogadoActivarForm}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Abogado_Activar.html', context)


def abogado_evaluacion_remota(request):

    usuario_actual=request.user.username
    abogado_actual = Abogados.objects.filter(username=usuario_actual)

    actas_remota = ActasRemotas.objects.filter(Q(sis_clasificacion__iexact="asignado_Abogado")|Q(sis_clasificacion__iexact="rechaza_encargado"), abogado_asignado=abogado_actual.first().rut)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    # actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Abogado_Evaluacion.html', context)

def abogado_desactivar_terreno(request, id):
    actas_terreno = ActasTerreno.objects.filter(global_id=id)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        if request.method == 'POST':
            GestionTerrenoAbogadoDesactivarForm = GestionTerrenoAbogadoDesactivar(request.POST, instance=acta)
            if GestionTerrenoAbogadoDesactivarForm.is_valid():
                acta = GestionTerrenoAbogadoDesactivarForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.abogado_resultado = GestionTerrenoAbogadoDesactivarForm.cleaned_data['abogado_resultado']
                acta.abogado_motivo_devolucion = GestionTerrenoAbogadoDesactivarForm.cleaned_data['abogado_motivo_devolucion']
                acta.sis_clasificacion = GestionTerrenoAbogadoDesactivarForm.cleaned_data['sis_clasificacion']
                acta.abogado_eleccion = GestionTerrenoAbogadoDesactivarForm.cleaned_data['abogado_eleccion']
                acta.abogado_presunto_infractor = GestionTerrenoAbogadoDesactivarForm.cleaned_data['abogado_presunto_infractor']
                acta.abogado_codigo_desactiva = GestionTerrenoAbogadoDesactivarForm.cleaned_data['abogado_codigo_desactiva']
                acta.abogado_obs = GestionTerrenoAbogadoDesactivarForm.cleaned_data['abogado_obs']
                acta.fecha_evaluacion_abogado = None
                if GestionTerrenoAbogadoDesactivarForm.cleaned_data['sis_clasificacion'] == 'abogado_desactivado':
                    acta.fecha_evaluacion_abogado = datetime.now()
                    WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                                 Usuario=request.user.username,
                                                 NuevaEtapa=GestionTerrenoAbogadoDesactivarForm.cleaned_data['sis_clasificacion'],
                                                 FechaCambio=datetime.now())

                acta.save(update_fields=['abogado_resultado','abogado_motivo_devolucion', 'sis_clasificacion', 'abogado_eleccion', 'abogado_presunto_infractor' , 'abogado_codigo_desactiva', 'abogado_obs','fecha_evaluacion_abogado'])
                return redirect('abogado_evaluacion_terreno')
            else:
                print(GestionTerrenoAbogadoDesactivar.errors)
                return redirect('abogado_evaluacion_terreno')
        else:
            GestionTerrenoAbogadoDesactivarForm = GestionTerrenoAbogadoDesactivar(instance=acta)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'GestionTerrenoAbogadoDesactivarForm': GestionTerrenoAbogadoDesactivarForm}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Abogado_Desactivar.html', context)

def abogado_activar_remota(request, id):
    actas_remota = ActasRemotas.objects.filter(global_id=id)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        try:
            latest_token = Tokens.objects.latest('id')
        except ObjectDoesNotExist:
            latest_token = None

        if request.method == 'POST':
            GestionRemotasAbogadoActivarForm = GestionRemotasAbogadoActivar(request.POST, instance=acta)
            if GestionRemotasAbogadoActivarForm.is_valid():
                acta = GestionRemotasAbogadoActivarForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.abogado_resultado = GestionRemotasAbogadoActivarForm.cleaned_data['abogado_resultado']
                acta.abogado_motivo_devolucion = GestionRemotasAbogadoActivarForm.cleaned_data[
                    'abogado_motivo_devolucion']
                acta.sis_clasificacion = GestionRemotasAbogadoActivarForm.cleaned_data['sis_clasificacion']
                acta.abogado_eleccion = GestionRemotasAbogadoActivarForm.cleaned_data['abogado_eleccion']
                acta.abogado_presunto_infractor = GestionRemotasAbogadoActivarForm.cleaned_data[
                    'abogado_presunto_infractor']
                acta.abogado_codigo_activa = GestionRemotasAbogadoActivarForm.cleaned_data['abogado_codigo_activa']
                acta.abogado_obs = GestionRemotasAbogadoActivarForm.cleaned_data['abogado_obs']
                acta.fecha_evaluacion_abogado = None
                if GestionRemotasAbogadoActivarForm.cleaned_data['sis_clasificacion'] == 'abogado_activado':
                    acta.fecha_evaluacion_abogado = datetime.now()
                    WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                                 Usuario=request.user.username,
                                                 NuevaEtapa=GestionRemotasAbogadoActivarForm.cleaned_data['sis_clasificacion'],
                                                 FechaCambio=datetime.now())
                acta.save(update_fields=['abogado_resultado', 'abogado_motivo_devolucion', 'sis_clasificacion',
                                         'abogado_eleccion', 'abogado_presunto_infractor', 'abogado_codigo_activa',
                                         'abogado_obs', 'fecha_evaluacion_abogado'])
                return redirect('abogado_evaluacion_remota')
            else:
                print(GestionRemotasAbogadoActivarForm.errors)
                return redirect('abogado_evaluacion_remota')
        else:
            GestionRemotasAbogadoActivarForm = GestionRemotasAbogadoActivar(instance=acta)

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'GestionRemotasAbogadoActivarForm': GestionRemotasAbogadoActivarForm}
    return render(request, 'GestionDenuncias/SGD2_Remota_Abogado_Activar.html', context)

def abogado_desactivar_remota(request, id):
    actas_remota = ActasRemotas.objects.filter(global_id=id)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        try:
            latest_token = Tokens.objects.latest('id')
        except ObjectDoesNotExist:
            latest_token = None

        if request.method == 'POST':
            GestionRemotasAbogadoDesactivarForm = GestionRemotasAbogadoDesactivar(request.POST, instance=acta)
            if GestionRemotasAbogadoDesactivarForm.is_valid():
                acta = GestionRemotasAbogadoDesactivarForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.abogado_resultado = GestionRemotasAbogadoDesactivarForm.cleaned_data['abogado_resultado']
                acta.abogado_motivo_devolucion = GestionRemotasAbogadoDesactivarForm.cleaned_data[
                    'abogado_motivo_devolucion']
                acta.sis_clasificacion = GestionRemotasAbogadoDesactivarForm.cleaned_data['sis_clasificacion']
                acta.abogado_eleccion = GestionRemotasAbogadoDesactivarForm.cleaned_data['abogado_eleccion']
                acta.abogado_presunto_infractor = GestionRemotasAbogadoDesactivarForm.cleaned_data[
                    'abogado_presunto_infractor']
                acta.abogado_codigo_desactiva = GestionRemotasAbogadoDesactivarForm.cleaned_data['abogado_codigo_desactiva']
                acta.abogado_obs = GestionRemotasAbogadoDesactivarForm.cleaned_data['abogado_obs']
                acta.fecha_evaluacion_abogado = None
                if GestionRemotasAbogadoDesactivarForm.cleaned_data['sis_clasificacion'] == 'abogado_desactivado':
                    acta.fecha_evaluacion_abogado = datetime.now()
                    WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                                 Usuario=request.user.username,
                                                 NuevaEtapa=GestionRemotasAbogadoDesactivarForm.cleaned_data['sis_clasificacion'],
                                                 FechaCambio=datetime.now())
                acta.save(update_fields=['abogado_resultado', 'abogado_motivo_devolucion', 'sis_clasificacion',
                                         'abogado_eleccion', 'abogado_presunto_infractor', 'abogado_codigo_desactiva',
                                         'abogado_obs','fecha_evaluacion_abogado'])
                return redirect('abogado_evaluacion_remota')
            else:
                print(GestionRemotasAbogadoDesactivarForm.errors)
                return redirect('abogado_evaluacion_remota')
        else:
            GestionRemotasAbogadoDesactivarForm = GestionRemotasAbogadoDesactivar(instance=acta)

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'GestionRemotasAbogadoDesactivarForm': GestionRemotasAbogadoDesactivarForm}
    return render(request, 'GestionDenuncias/SGD2_Remota_Abogado_Desactivar.html', context)

def abogado_activadas_terreno(request):

    usuario_actual=request.user.username
    abogado_actual = Abogados.objects.filter(username=usuario_actual)

    actas_terreno = ActasTerreno.objects.filter(sis_clasificacion="abogado_activado", abogado_asignado=abogado_actual.first().rut)

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Abogado_Activadas.html', context)

def abogado_activadas_terreno_gestiones(request, id):
    actas_terreno = ActasTerreno.objects.filter(global_id=id)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        if request.method == 'POST':
            GestionTerrenoAbogadoActivarGestionesForm = GestionTerrenoAbogadoActivarGestiones(request.POST, instance=acta)
            if GestionTerrenoAbogadoActivarGestionesForm.is_valid():
                acta = GestionTerrenoAbogadoActivarGestionesForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía


                acta.sis_clasificacion = GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['sis_clasificacion']
                acta.abogado_eleccion = GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['abogado_eleccion']
                acta.abogado_presunto_infractor = GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['abogado_presunto_infractor']
                acta.abogado_codigo_activa = GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['abogado_codigo_activa']
                acta.abogado_obs = GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['abogado_obs']

                acta.abogado_folio = GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['abogado_folio']
                acta.abogado_obs_finales = GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['abogado_obs_finales']
                acta.abogado_resultado_final = GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['abogado_resultado_final']
                acta.fecha_envio_abogado = None
                if GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['sis_clasificacion'] in [
                    'abogado_con_infraccion', 'abogado_sin_infraccion']:
                    acta.fecha_envio_abogado = datetime.now()
                    WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                                 Usuario=request.user.username,
                                                 NuevaEtapa=GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['sis_clasificacion'],
                                                 FechaCambio=datetime.now())
                acta.save(update_fields=['sis_clasificacion', 'abogado_eleccion', 'abogado_presunto_infractor' , 'abogado_codigo_activa', 'abogado_obs','abogado_folio','abogado_obs_finales','abogado_resultado_final','fecha_envio_abogado'])
                print(GestionTerrenoAbogadoActivarGestionesForm.cleaned_data['sis_clasificacion'])
                return redirect('abogado_activadas_terreno')
            else:
                print(GestionTerrenoAbogadoActivarGestionesForm.errors)
                print(acta.sis_clasificacion)
                return redirect('abogado_activadas_terreno')

        else:
            GestionTerrenoAbogadoActivarGestionesForm = GestionTerrenoAbogadoActivarGestiones(instance=acta)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'GestionTerrenoAbogadoActivarGestionesForm': GestionTerrenoAbogadoActivarGestionesForm}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Abogado_Activadas_Gestiones.html', context)

def abogado_activadas_remotas(request):

    usuario_actual=request.user.username
    abogado_actual = Abogados.objects.filter(username=usuario_actual)

    actas_remota = ActasRemotas.objects.filter(sis_clasificacion="abogado_activado", abogado_asignado=abogado_actual.first().rut)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    # actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Abogado_Activadas.html', context)

def abogado_activadas_remotas_gestiones(request, id):
    actas_remota = ActasRemotas.objects.filter(global_id=id)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        try:
            latest_token = Tokens.objects.latest('id')
        except ObjectDoesNotExist:
            latest_token = None

        if request.method == 'POST':
            GestionRemotasAbogadoActivarGestionesForm = GestionRemotasAbogadoActivarGestiones(request.POST, instance=acta)
            if GestionRemotasAbogadoActivarGestionesForm.is_valid():
                acta = GestionRemotasAbogadoActivarGestionesForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.sis_clasificacion = GestionRemotasAbogadoActivarGestionesForm.cleaned_data['sis_clasificacion']
                acta.abogado_eleccion = GestionRemotasAbogadoActivarGestionesForm.cleaned_data['abogado_eleccion']
                acta.abogado_presunto_infractor = GestionRemotasAbogadoActivarGestionesForm.cleaned_data['abogado_presunto_infractor']
                acta.abogado_codigo_activa = GestionRemotasAbogadoActivarGestionesForm.cleaned_data['abogado_codigo_activa']
                acta.abogado_obs = GestionRemotasAbogadoActivarGestionesForm.cleaned_data['abogado_obs']

                acta.abogado_folio = GestionRemotasAbogadoActivarGestionesForm.cleaned_data['abogado_folio']
                acta.abogado_obs_finales = GestionRemotasAbogadoActivarGestionesForm.cleaned_data['abogado_obs_finales']
                acta.abogado_resultado_final = GestionRemotasAbogadoActivarGestionesForm.cleaned_data['abogado_resultado_final']
                acta.fecha_envio_abogado = None
                if GestionRemotasAbogadoActivarGestionesForm.cleaned_data['sis_clasificacion'] in [
                    'abogado_con_infraccion', 'abogado_sin_infraccion']:
                    acta.fecha_envio_abogado = datetime.now()
                    WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                                 Usuario=request.user.username,
                                                 NuevaEtapa=GestionRemotasAbogadoActivarGestionesForm.cleaned_data['sis_clasificacion'],
                                                 FechaCambio=datetime.now())
                acta.save(update_fields=['sis_clasificacion', 'abogado_eleccion', 'abogado_presunto_infractor' , 'abogado_codigo_activa', 'abogado_obs','abogado_folio','abogado_obs_finales','abogado_resultado_final', 'fecha_envio_abogado'])
                return redirect('abogado_activadas_remotas')
            else:
                print(GestionRemotasAbogadoActivarGestionesForm.errors)
                return redirect('abogado_activadas_remotas')
        else:
            GestionRemotasAbogadoActivarGestionesForm = GestionRemotasAbogadoActivarGestiones(instance=acta)

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'GestionRemotasAbogadoActivarGestionesForm': GestionRemotasAbogadoActivarGestionesForm}
    return render(request, 'GestionDenuncias/SGD2_Remota_Abogado_Activadas_Gestiones.html', context)


def efr_remota_devuelto_abogado(request):
    try:
        revisor = EFRDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except EFRDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_remota = ActasRemotas.objects.filter(sis_clasificacion="abogado_devuelto", region=region_usuario)
    else:
        actas_remota = []

    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    # actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_EFR_Devuelto_Abogado.html', context)


def efr_remota_devuelta_abogado_gestiones(request, id):
    actas_remota = ActasRemotas.objects.filter(global_id=id)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        try:
            latest_token = Tokens.objects.latest('id')
        except ObjectDoesNotExist:
            latest_token = None

        if request.method == 'POST':
            GestionRemotasEFRForm = GestionRemotasEFR(request.POST, instance=acta)
            if GestionRemotasEFRForm.is_valid():
                acta = GestionRemotasEFRForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.sis_resultado_efr = GestionRemotasEFRForm.cleaned_data['sis_resultado_efr']
                acta.sis_motivo_rechazo = GestionRemotasEFRForm.cleaned_data['sis_motivo_rechazo']
                acta.sis_clasificacion = GestionRemotasEFRForm.cleaned_data['sis_clasificacion']
                WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                             Usuario=request.user.username,
                                             NuevaEtapa=GestionRemotasEFRForm.cleaned_data['sis_clasificacion'],
                                             FechaCambio=datetime.now())

                acta.save(update_fields=['sis_resultado_efr', 'sis_motivo_rechazo', 'sis_clasificacion'])

                return redirect('efr_remota_devuelto_abogado')
            else:
                print(GestionRemotasEFRForm.errors)
                return redirect('efr_remota_devuelto_abogado')
        else:
            GestionRemotasEFRForm = GestionRemotasEFR(instance=acta)

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'GestionRemotasEFRForm': GestionRemotasEFRForm}
    return render(request, 'GestionDenuncias/SGD2_Remota_EFR_Devuelto_Abogado_Gestiones.html', context)

def efr_terreno_devuelto_abogado(request):
    try:
        revisor = EFRDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except EFRDR.DoesNotExist:
        # Manejo de error en caso de que no se encuentre la región del usuario
        region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_terreno = ActasTerreno.objects.filter(sis_clasificacion="abogado_devuelto", region=region_usuario)
    else:
        actas_terreno = []

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_EFR_Devuelto_Abogado.html', context)


def efr_terreno_devuelta_abogado_gestiones(request, id):
    actas_terreno = ActasTerreno.objects.filter(global_id=id)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        if request.method == 'POST':
            GestionTerrenoEFRForm = GestionTerrenoEFR(request.POST, instance=acta)
            if GestionTerrenoEFRForm.is_valid():
                acta = GestionTerrenoEFRForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.sis_resultado_efr = GestionTerrenoEFRForm.cleaned_data['sis_resultado_efr']
                acta.sis_motivo_rechazo = GestionTerrenoEFRForm.cleaned_data['sis_motivo_rechazo']
                acta.sis_clasificacion = GestionTerrenoEFRForm.cleaned_data['sis_clasificacion']
                acta.save(update_fields=['sis_resultado_efr','sis_motivo_rechazo', 'sis_clasificacion' ])
                WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                             Usuario=request.user.username,
                                             NuevaEtapa=GestionTerrenoEFRForm.cleaned_data['sis_clasificacion'],
                                             FechaCambio=datetime.now())
                return redirect('efr_terreno_devuelto_abogado')
            else:
                print(GestionTerrenoEFRForm.errors)
                return redirect('efr_terreno_devuelto_abogado')
        else:
            GestionTerrenoEFRForm = GestionTerrenoEFR(instance=acta)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'GestionTerrenoEFRForm': GestionTerrenoEFRForm}
    return render(request, 'GestionDenuncias/SGD2_Terreno_EFR_Devuelto_Abogado_Gestiones.html', context)




def insert_token(request):
    if request.method == 'POST':
        form = TokenForm(request.POST)
        if form.is_valid():
            token_value = form.cleaned_data['token_value']
            # Asegúrate de limpiar y validar este valor adecuadamente
            with connection.cursor() as cursor:
                cursor.execute("INSERT INTO public.\"GestionDenuncias_tokens\" (\"Token\", \"Fecha\") VALUES (%s, NOW())", [token_value])
            # Redirecciona o muestra un mensaje de éxito
    else:
        form = TokenForm()

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    return render(request, 'GestionDenuncias/SGD2_ingresa_Token.html', {'latest_token': latest_token,'form': form})


def encargado_terreno_revision(request):

    actas_terreno = ActasTerreno.objects.filter(Q(sis_clasificacion__iexact="abogado_con_infraccion")|Q(sis_clasificacion__iexact="abogado_sin_infraccion"))

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Encargado.html', context)

def encargado_terreno_revision_gestiones(request, id):
    actas_terreno = ActasTerreno.objects.filter(global_id=id)
    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        if request.method == 'POST':
            GestionTerrenoEncargadoGestionesForm = GestionTerrenoEncargadoGestiones(request.POST, instance=acta)
            if GestionTerrenoEncargadoGestionesForm.is_valid():
                acta = GestionTerrenoEncargadoGestionesForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.sis_encargado_resultado = GestionTerrenoEncargadoGestionesForm.cleaned_data['sis_encargado_resultado']
                acta.sis_motivo_rechazo_encargado = GestionTerrenoEncargadoGestionesForm.cleaned_data['sis_motivo_rechazo_encargado']
                acta.sis_clasificacion = GestionTerrenoEncargadoGestionesForm.cleaned_data['sis_clasificacion']
                WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                             Usuario=request.user.username,
                                             NuevaEtapa=GestionTerrenoEncargadoGestionesForm.cleaned_data['sis_clasificacion'],
                                             FechaCambio=datetime.now())
                acta.save(update_fields=['sis_encargado_resultado','sis_motivo_rechazo_encargado', 'sis_clasificacion' ])
                return redirect('encargado_terreno_revision')
            else:
                print(GestionTerrenoEncargadoGestionesForm.errors)
                return redirect('encargado_terreno_revision')
        else:
            GestionTerrenoEncargadoGestionesForm = GestionTerrenoEncargadoGestiones(instance=acta)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno, 'GestionTerrenoEncargadoGestionesForm': GestionTerrenoEncargadoGestionesForm}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Encargado_Gestiones.html', context)


def encargado_remota_revision(request):

    actas_remota = ActasRemotas.objects.filter(Q(sis_clasificacion__iexact="abogado_con_infraccion")|Q(sis_clasificacion__iexact="abogado_sin_infraccion"))


    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    # actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Encargado.html', context)

def encargado_remota_revision_gestiones(request, id):
    actas_remota = ActasRemotas.objects.filter(global_id=id)
    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

        try:
            latest_token = Tokens.objects.latest('id')
        except ObjectDoesNotExist:
            latest_token = None

        if request.method == 'POST':
            GestionRemotasEncargadoGestionesForm = GestionRemotasEncargadoGestiones(request.POST, instance=acta)
            if GestionRemotasEncargadoGestionesForm.is_valid():
                acta = GestionRemotasEncargadoGestionesForm.save(commit=False)  # Esto no guarda el objeto en la base de datos todavía
                acta.sis_clasificacion = GestionRemotasEncargadoGestionesForm.cleaned_data['sis_clasificacion']
                acta.sis_encargado_resultado = GestionRemotasEncargadoGestionesForm.cleaned_data['sis_encargado_resultado']
                acta.sis_motivo_rechazo_encargado = GestionRemotasEncargadoGestionesForm.cleaned_data['sis_motivo_rechazo_encargado']
                WorkflowActas.objects.create(GlobalID=str(acta.global_id),
                                             Usuario=request.user.username,
                                             NuevaEtapa=GestionRemotasEncargadoGestionesForm.cleaned_data['sis_clasificacion'],
                                             FechaCambio=datetime.now())

                acta.save(update_fields=['sis_clasificacion', 'sis_encargado_resultado', 'sis_motivo_rechazo_encargado' ])
                return redirect('encargado_remota_revision')
            else:
                print(GestionRemotasEncargadoGestionesForm.errors)
                return redirect('encargado_remota_revision')
        else:
            GestionRemotasEncargadoGestionesForm = GestionRemotasEncargadoGestiones(instance=acta)

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota, 'GestionRemotasEncargadoGestionesForm': GestionRemotasEncargadoGestionesForm}
    return render(request, 'GestionDenuncias/SGD2_Remota_Encargado_Gestiones.html', context)

def encargado_remota_despacho(request):

    actas_remota = ActasRemotas.objects.filter(sis_encargado_resultado="Acepta")


    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    # actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Encargado_Despacho.html', context)


def encargado_terreno_despacho(request):

    actas_terreno = ActasTerreno.objects.filter(sis_encargado_resultado="Acepta")

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Encargado_Despacho.html', context)

def dr_terreno_reporte(request):
    try:
        # Intenta obtener el revisor del modelo RevisoresDR
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        try:
            # Si no se encuentra en RevisoresDR, intenta obtenerlo del modelo EFRDR
            revisor = EFRDR.objects.get(id_usuario__username=request.user)
            region_usuario = revisor.Region
        except EFRDR.DoesNotExist:
            # Manejo de error en caso de que no se encuentre la región del usuario en ninguno de los modelos
            region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_terreno = ActasTerreno.objects.filter(region=region_usuario).exclude(sis_clasificacion__in=['Pendiente', 'archivo_terreno'])
    else:
        actas_terreno = []

    # Aca en icontains pongo el filtro con el metodo icontains que es un like

    for acta in actas_terreno:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 =  increment_url_numbers(acta.evidencia_fotografica)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None

    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_terreno': actas_terreno}
    return render(request, 'GestionDenuncias/SGD2_Terreno_Reporte_DR.html', context)

def dr_remota_reporte(request):
    try:
        # Intenta obtener el revisor del modelo RevisoresDR
        revisor = RevisoresDR.objects.get(id_usuario__username=request.user)
        region_usuario = revisor.Region
    except RevisoresDR.DoesNotExist:
        try:
            # Si no se encuentra en RevisoresDR, intenta obtenerlo del modelo EFRDR
            revisor = EFRDR.objects.get(id_usuario__username=request.user)
            region_usuario = revisor.Region
        except EFRDR.DoesNotExist:
            # Manejo de error en caso de que no se encuentre la región del usuario en ninguno de los modelos
            region_usuario = None

    if region_usuario:
        # Filtra las ActasTerreno basadas en la región del usuario
        actas_remota = ActasRemotas.objects.filter(region=region_usuario).exclude(sis_clasificacion__in=['Pendiente', 'archivo_remota'])
    else:
        actas_remota = []
    # Aca en icontains pongo el filtro con el metodo icontains que es un like

    for acta in actas_remota:
        # Asumiendo que tu valor epoch está en milisegundos. Si está en segundos, omite la división por 1000.
        local_date = datetime.utcfromtimestamp(int(acta.creation_date) / 1000)
        acta.adjunto2 = increment_url_numbers(acta.medios_respaldo_adjunto)
        acta.adjunto3 = increment_url_numbers(acta.adjunto2)
        acta.adjunto4 = increment_url_numbers(acta.adjunto3)
        acta.adjunto5 = increment_url_numbers(acta.adjunto4)
        acta.adjunto6 = increment_url_numbers(acta.adjunto5)

        acta.audio2 = increment_url_numbers(acta.ingrese_audios)
        acta.audio3 = increment_url_numbers(acta.audio2)
        acta.audio4 = increment_url_numbers(acta.audio3)
        acta.audio5 = increment_url_numbers(acta.audio4)
        acta.audio6 = increment_url_numbers(acta.audio5)
        # Restar 3 horas
        acta.fecha = local_date - timedelta(hours=3)

    try:
        latest_token = Tokens.objects.latest('id')
    except ObjectDoesNotExist:
        latest_token = None



    #actas_remotas = ActasRemotas.objects.filter(sis_clasificacion="Pendiente")
    context = {'latest_token': latest_token, 'actas_remota': actas_remota}
    return render(request, 'GestionDenuncias/SGD2_Remota_Reporte_DR.html', context)
